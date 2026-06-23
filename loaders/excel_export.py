import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import OUTPUT_FILENAME

def export_to_excel(df_campaigns, df_ads):
    """
    Exporta los datasets a un fichero Excel con dos tabs:
    - campaigns: métricas agregadas por campaña
    - ads: métricas por anuncio/creativo
    """

    if df_campaigns.empty and df_ads.empty:
        print("⚠️ No hay datos para exportar")
        return

    # Columnas y orden para cada tab
    columns_campaigns = [
        "campaign_name", "goal",
        "start_date", "end_date", "budget", "currency",
        "impressions", "reach", "clicks", "landing_page_clicks",
        "spend", "CTR", "CPM", "CPC",
        "video_views", "video_view_rate", "CPV",
        "video_25pct", "video_50pct", "video_75pct",
        "video_completions", "completion_rate",
        "engagements", "engagement_rate",
        "likes", "comments", "shares", "follows",
    ]

    columns_ads = [
        "campaign_id", "creative_id",
        "impressions", "reach", "clicks",
        "spend", "CTR", "CPM", "CPC",
        "video_views", "video_view_rate", "CPV",
        "video_25pct", "video_50pct", "video_75pct",
        "video_completions", "completion_rate",
        "engagements", "engagement_rate",
        "likes", "comments", "shares", "follows",
        "landing_page_clicks",
    ]

    # Filtrar solo columnas que existen en el dataframe
    cols_camp = [c for c in columns_campaigns if c in df_campaigns.columns]
    cols_ads  = [c for c in columns_ads if c in df_ads.columns]

    with pd.ExcelWriter(OUTPUT_FILENAME, engine="openpyxl") as writer:

        # Escribir tabs
        if not df_campaigns.empty:
            df_campaigns[cols_camp].to_excel(writer, sheet_name="campaigns", index=False)

        if not df_ads.empty:
            df_ads[cols_ads].to_excel(writer, sheet_name="ads", index=False)

        workbook = writer.book

        # Estilos
        header_font    = Font(bold=True, color="FFFFFF", size=11)
        header_fill    = PatternFill("solid", fgColor="0A66C2")  # Azul LinkedIn
        header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align     = Alignment(horizontal="center", vertical="center")
        thin_border    = Border(
            left   = Side(style="thin", color="D3D3D3"),
            right  = Side(style="thin", color="D3D3D3"),
            top    = Side(style="thin", color="D3D3D3"),
            bottom = Side(style="thin", color="D3D3D3"),
        )

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            # Formato de cabeceras
            for cell in ws[1]:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = header_align
                cell.border    = thin_border

            # Formato de celdas de datos
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = cell_align
                    cell.border    = thin_border

            # Ajustar ancho de columnas automáticamente
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

            # Fijar primera fila
            ws.freeze_panes = "A2"

    print(f"✅ Excel exportado correctamente: {OUTPUT_FILENAME}")
    print(f"   Tabs: {', '.join(workbook.sheetnames)}")

if __name__ == "__main__":
    print("Este módulo se ejecuta desde main.py")